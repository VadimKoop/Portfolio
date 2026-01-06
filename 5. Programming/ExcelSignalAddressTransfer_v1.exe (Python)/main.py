"""
main.py — logic for transferring Obj.Addr.^ between Excel files.

UI (app_ui.py / ui_app.py) calls:
run_transfer(source_file, target_file, output_folder)

Additionally:
- Two report sheets are added to the output file:
* "Written" — a list of signals for which the address was found and written
* "Not found" — a list of signals for which the address was not found

Buttons (exactly 5):
- Old file with signals -> source_file
- New file without signals -> target_file
- Output folder -> output_folder
- Cancel
- Execute
"""

from __future__ import annotations

import openpyxl
from openpyxl.utils import get_column_letter


def find_header_and_cols(ws):
    """
    Find:
      - header_row: row index where column 1 == 'Signal'
      - sig_col: column index where header == 'Signal name'
      - obj_col: column index where header contains 'Obj. Addr'
    """
    header_row = None
    sig_col = None
    obj_col = None

    for r in range(1, ws.max_row + 1):
        v1 = ws.cell(row=r, column=1).value
        if isinstance(v1, str) and v1.strip() == "Signal":
            header_row = r
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str):
                    vv = v.strip()
                    if vv == "Signal name":
                        sig_col = c
                    if "Obj. Addr" in vv:
                        obj_col = c
            break

    return header_row, sig_col, obj_col


def find_first_empty_header_col(ws):
    """
    In the row where column 1 == 'Signal', find the first column
    where the header cell is empty (None). Return (header_row, col_index).
    """
    header_row, _, _ = find_header_and_cols(ws)
    if header_row is None:
        raise ValueError("Header row ('Signal') not found.")

    max_col = ws.max_column
    for c in range(1, max_col + 2):
        if ws.cell(row=header_row, column=c).value is None:
            return header_row, c

    return header_row, max_col + 1


def build_mapping(ws):
    """
    Build a mapping from the source sheet:

        (Database, Signal name) -> Obj. Addr. ^

    - Database is taken from rows where col1 == 'Database', value in col4.
    - Signal name is taken from the 'Signal name' column in data rows.
    - Obj. Addr. ^ is taken from the 'Obj. Addr. ^' column in data rows.
    """
    _, sig_col, obj_col = find_header_and_cols(ws)
    if sig_col is None or obj_col is None:
        raise ValueError("Could not find 'Signal name' or 'Obj. Addr. ^' columns in source sheet.")

    mapping = {}
    current_db = None

    for r in range(1, ws.max_row + 1):
        v1 = ws.cell(row=r, column=1).value

        # Database row
        if isinstance(v1, str) and v1.strip() == "Database":
            current_db = ws.cell(row=r, column=4).value
            continue

        # Header row inside the block
        if isinstance(v1, str) and v1.strip() == "Signal":
            continue

        sig = ws.cell(row=r, column=sig_col).value
        if sig is None or str(sig).strip() == "":
            continue
        if current_db is None:
            continue

        key = (str(current_db).strip(), str(sig).strip())
        obj = ws.cell(row=r, column=obj_col).value

        if obj is not None and str(obj).strip() != "":
            mapping[key] = obj

    return mapping


def transfer_obj_addr_between_sheets(ws_src, ws_dst, sheet_name: str, written_rows: list, not_found_rows: list):
    """
    Transfer Obj. Addr. ^ from ws_src to ws_dst for one sheet.
    Matching by (Database, Signal name).

    In the target sheet:
      - create first empty header column named 'Obj. Addr. ^'
      - write value or 'not found'

    Also collects rows for report sheets:
      - written_rows:  (sheet, database, signal_name, obj_addr)
      - not_found_rows:(sheet, database, signal_name)
    """
    mapping = build_mapping(ws_src)

    header_row_dst, empty_col = find_first_empty_header_col(ws_dst)
    ws_dst.cell(row=header_row_dst, column=empty_col).value = "Obj. Addr. ^"

    _, sig_col_dst, _ = find_header_and_cols(ws_dst)
    if sig_col_dst is None:
        raise ValueError(f"Could not find 'Signal name' column in target sheet {sheet_name}.")

    current_db = None
    written_count = 0
    not_found_count = 0

    for r in range(1, ws_dst.max_row + 1):
        v1 = ws_dst.cell(row=r, column=1).value

        # Database row
        if isinstance(v1, str) and v1.strip() == "Database":
            current_db = ws_dst.cell(row=r, column=4).value
            continue

        # Header row inside the block
        if isinstance(v1, str) and v1.strip() == "Signal":
            continue

        sig = ws_dst.cell(row=r, column=sig_col_dst).value
        if sig is None or str(sig).strip() == "":
            continue
        if current_db is None:
            continue

        db_str = str(current_db).strip()
        sig_str = str(sig).strip()
        key = (db_str, sig_str)

        obj = mapping.get(key)

        if obj is not None and str(obj).strip() != "":
            ws_dst.cell(row=r, column=empty_col).value = obj
            written_count += 1
            written_rows.append((sheet_name, db_str, sig_str, obj))
        else:
            ws_dst.cell(row=r, column=empty_col).value = "not found"
            not_found_count += 1
            not_found_rows.append((sheet_name, db_str, sig_str))

    return written_count, not_found_count


def _replace_sheet(wb, title: str):
    if title in wb.sheetnames:
        ws_old = wb[title]
        wb.remove(ws_old)
    return wb.create_sheet(title)


def _write_table(ws, headers, rows):
    ws.append(list(headers))
    for row in rows:
        ws.append(list(row))

    ws.freeze_panes = "A2"

    # Auto-filter range
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row >= 1 and max_col >= 1:
        last_col_letter = get_column_letter(max_col)
        ws.auto_filter.ref = f"A1:{last_col_letter}{max_row}"

    # Basic column width (simple heuristic)
    for c in range(1, ws.max_column + 1):
        col_letter = get_column_letter(c)
        ws.column_dimensions[col_letter].width = 22


def add_report_sheets(wb_dst, written_rows: list, not_found_rows: list):
    """
    Adds/overwrites two sheets in the target workbook:
      - "Written"
      - "Not found"
    """
    ws_written = _replace_sheet(wb_dst, "Written")
    _write_table(ws_written, ["Sheet", "Database", "Signal name", "Obj. Addr. ^"], written_rows)

    ws_nf = _replace_sheet(wb_dst, "Not found")
    _write_table(ws_nf, ["Sheet", "Database", "Signal name"], not_found_rows)


def run_transfer(source_file: str, target_file: str, output_folder: str, sheets=("AI", "BI", "BO")):
    """
    Runs the transfer for the given files and returns (total_written, total_not_found).
    Also writes two report sheets into the output workbook.
    """
    wb_src = openpyxl.load_workbook(source_file, data_only=True)
    wb_dst = openpyxl.load_workbook(target_file)

    total_written = 0
    total_not_found = 0

    written_rows = []
    not_found_rows = []

    for sheet in sheets:
        if sheet not in wb_src.sheetnames:
            raise ValueError(f"Source workbook has no sheet '{sheet}'. Found: {wb_src.sheetnames}")
        if sheet not in wb_dst.sheetnames:
            raise ValueError(f"Target workbook has no sheet '{sheet}'. Found: {wb_dst.sheetnames}")

        ws_src = wb_src[sheet]
        ws_dst = wb_dst[sheet]
        written, not_found = transfer_obj_addr_between_sheets(ws_src, ws_dst, sheet, written_rows, not_found_rows)
        total_written += written
        total_not_found += not_found

    # Add report sheets into output workbook
    add_report_sheets(wb_dst, written_rows, not_found_rows)

    wb_dst.save(output_folder)
    return total_written, total_not_found
